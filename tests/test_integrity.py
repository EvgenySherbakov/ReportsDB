"""Проверки целостности. См. docs/TZ.md, раздел 12.

Запуск: pytest -q   (перед этим: python -m reportsdb sample)
"""

from __future__ import annotations

import sys
from pathlib import Path

import duckdb
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "app"))

from reportsdb.config import (  # noqa: E402
    Mapping,
    SectionConfig,
    load_mapping,
    resolve_columns,
)
from reportsdb.config import SCHEMA_VERSION  # noqa: E402
from reportsdb.etl import build, clear  # noqa: E402
from reportsdb.normalize import (  # noqa: E402
    join_folders,
    normalise_catalog_path,
    parse_bool,
    parse_table_list,
    parse_table_ref,
)
from reportsdb.sample_data import generate  # noqa: E402

VIEWS = [
    "v_report_footprint",
    "v_table_criticality",
    "v_report_cost_value",
    "v_decommission_candidates",
    "v_catalog_overview",
    "v_schema_overview",
    "v_network_overview",
    "v_report_duration",
    "v_report_overlap",
    "v_report_plant_twin",
    "v_tables_catalog",
    "v_report_table_size",
    "v_rc_report_tables",
    "v_rc_report_routines",
    "v_rc_report_usage",
    "v_rc_report_retention",
    "v_rc_summary",
    "v_report_tables_summary",
]

# Каждое поле, приходящее из исходных файлов, обязано быть видно в аналитике —
# иначе колонку загрузили впустую. Проверяется тестом ниже.
FIELDS_IN_ANALYTICS = [
    "report_no", "report_name", "network", "plant",
    "folder_l1", "folder_l2", "folder_l3", "uses_view",
    "exec_count", "avg_duration_sec", "total_duration_sec",
    "percent_of_total", "exclusive_pct_of_db", "segment_count",
    "total_mb", "schema_name", "table_name",
    "object_kind", "kind_source", "schema_source", "retention_days", "retention_band",
    "usage_band",
    "tables_total_mb", "tables_exclusive_mb", "table_names",
]


@pytest.fixture(scope="module")
def paths(tmp_path_factory):
    return generate()


@pytest.fixture(scope="module")
def full_db(paths, tmp_path_factory):
    """БД со всеми тремя источниками."""
    mapping = load_mapping()
    mapping.table_sizes.file = str(paths["sizes"])
    mapping.report_usage.file = str(paths["usage"])
    db = tmp_path_factory.mktemp("db") / "full.duckdb"
    build(paths["reports"], db, mapping)
    con = duckdb.connect(str(db), read_only=True)
    yield con
    con.close()


@pytest.fixture(scope="module")
def bare_db(paths, tmp_path_factory):
    """БД только с каталогом отчётов — факты пусты."""
    base = load_mapping()
    mapping = Mapping(
        reports=base.reports,
        table_sizes=SectionConfig(),
        report_usage=SectionConfig(),
    )
    db = tmp_path_factory.mktemp("db") / "bare.duckdb"
    build(paths["reports"], db, mapping)
    con = duckdb.connect(str(db), read_only=True)
    yield con
    con.close()


# --- Нормализация ---------------------------------------------------------

@pytest.mark.parametrize(
    "raw,expected",
    [
        ("dbo.Orders", ("dbo", "Orders", True)),
        ("[dbo].[Orders]", ("dbo", "Orders", True)),
        ("  fin.Invoice  ", ("fin", "Invoice", True)),
        ("ReportDW.dbo.Orders", ("dbo", "Orders", True)),  # берём два последних
        ("TempStaging", ("(unknown)", "TempStaging", False)),
        ("", None),
        ("   ", None),
        (".", None),
    ],
)
def test_parse_table_ref(raw, expected):
    assert parse_table_ref(raw) == expected


def test_parse_table_list_dedupes_case_insensitively():
    result = parse_table_list("dbo.Orders; DBO.ORDERS ;dbo.Customers;;")
    assert len(result) == 2


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("/Finance/Monthly", "/Finance/Monthly"),
        ("\\Finance\\Monthly\\", "/Finance/Monthly"),
        ("Finance//Monthly", "/Finance/Monthly"),
        ("", "/"),
        (None, "/"),
        ("/", "/"),
    ],
)
def test_normalise_catalog_path(raw, expected):
    assert normalise_catalog_path(raw) == expected


def test_real_headers_from_customer_file_resolve():
    """Фактические заголовки файла отчётов сопоставляются без правок конфига.

    «Наименование отчета» пишется без «ё» — сопоставление это учитывает.
    """
    headers = [
        "№", "ТС", "Завод", "Каталог 1-го уровня", "Каталог 2-го уровня",
        "Каталог 3-го уровня", "Наименование отчета", "Используется view",
        "Таблицы источники данных", "Ср. дл. (сек)", "Кол-во обращений",
    ]
    resolved = resolve_columns(headers, load_mapping().reports.columns)
    assert resolved["report_no"] == "№"
    assert resolved["network"] == "ТС"
    assert resolved["plant"] == "Завод"
    assert resolved["folder_l1"] == "Каталог 1-го уровня"
    assert resolved["folder_l2"] == "Каталог 2-го уровня"
    assert resolved["folder_l3"] == "Каталог 3-го уровня"
    assert resolved["report_name"] == "Наименование отчета"
    assert resolved["uses_view"] == "Используется view"
    assert resolved["source_tables"] == "Таблицы источники данных"
    assert resolved["avg_duration_sec"] == "Ср. дл. (сек)"
    assert resolved["exec_count"] == "Кол-во обращений"


def test_segment_export_headers_resolve():
    """Заголовки выгрузки сегментов БД сопоставляются без правок конфига."""
    headers = ["№", "OWNER", "SEGMENT_NAME", "SEGMENT_TYPE", "SIZE_MB",
               "PERCENT_OF_TOTAL", "PERCENT_OF_SCHEMA"]
    resolved = resolve_columns(headers, load_mapping().table_sizes.columns)
    assert resolved["schema_name"] == "OWNER"
    assert resolved["table_name"] == "SEGMENT_NAME"
    assert resolved["segment_type"] == "SEGMENT_TYPE"
    assert resolved["total_mb"] == "SIZE_MB"
    assert resolved["percent_of_total"] == "PERCENT_OF_TOTAL"


@pytest.mark.parametrize(
    "levels,expected",
    [
        (("Финансы", "Месяц", "Продажи"), "/Финансы/Месяц/Продажи"),
        (("Финансы", "", ""), "/Финансы"),
        (("Финансы", None, "Продажи"), "/Финансы/Продажи"),  # пустой уровень пропускается
        ((None, None, None), "/"),
        (("A/B", "C", None), "/A-B/C"),  # слэш внутри уровня не ломает путь
    ],
)
def test_join_folders(levels, expected):
    assert join_folders(*levels) == expected


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("да", True), ("Да", True), ("ДА", True), ("yes", True), ("1", True), ("+", True),
        ("нет", False), ("Нет", False), ("no", False), ("0", False), ("-", False),
        ("", None), (None, None), ("может быть", None),  # непонятное — не False
    ],
)
def test_parse_bool(raw, expected):
    assert parse_bool(raw) is expected


def test_uses_view_and_org_dimensions_loaded(full_db):
    row = full_db.execute(
        "SELECT COUNT(*) FILTER (WHERE network IS NOT NULL), "
        "       COUNT(*) FILTER (WHERE plant IS NOT NULL), "
        "       COUNT(*) FILTER (WHERE uses_view IS NOT NULL) FROM dim_report"
    ).fetchone()
    assert all(v > 0 for v in row), "ТС, завод и признак view должны загружаться"


def test_same_report_name_kept_per_plant(full_db):
    """Одно имя отчёта у разных заводов — разные строки, а не одна."""
    collapsed = full_db.execute(
        "SELECT COUNT(*) FROM (SELECT report_name FROM dim_report "
        "GROUP BY report_name HAVING COUNT(DISTINCT plant) > 1 AND COUNT(*) = 1)"
    ).fetchone()[0]
    assert collapsed == 0


def test_segments_are_summed_not_overwritten(full_db):
    """Секции одной таблицы складываются, а не затирают друг друга."""
    multi = full_db.execute(
        "SELECT COUNT(*) FROM fact_table_size WHERE segment_count > 1"
    ).fetchone()[0]
    assert multi > 0, "в демо-данных есть секционированные таблицы"
    bad = full_db.execute(
        "SELECT COUNT(*) FROM fact_table_size WHERE segment_count > 1 AND total_mb IS NULL"
    ).fetchone()[0]
    assert bad == 0


def test_index_segments_are_not_counted_as_tables(full_db):
    """Индексные сегменты не должны попадать в размеры таблиц."""
    leaked = full_db.execute(
        "SELECT COUNT(*) FROM dim_table WHERE table_name ILIKE 'IDX!_%' ESCAPE '!'"
    ).fetchone()[0]
    assert leaked == 0


def _split_by_plant(path: Path, out_dir: Path, prefix: str) -> list[Path]:
    """Режет файл по заводам — имитирует выгрузку «файл на завод»."""
    import pandas as pd

    df = pd.read_excel(path, dtype=str)
    made = []
    for plant in sorted(df["Завод"].dropna().unique()):
        target = out_dir / f"{prefix}_{plant.replace(' ', '')}.xlsx"
        df[df["Завод"] == plant].to_excel(target, index=False)
        made.append(target)
    return made


def test_files_per_plant_load_same_as_one_file(paths, tmp_path_factory):
    """Разрезанные по заводам файлы дают ту же базу, что и один общий файл.

    Данные приходят по файлу на завод, и склейка не должна ни терять строки,
    ни складывать размеры разных заводов в один.
    """
    split_dir = tmp_path_factory.mktemp("split")
    report_files = _split_by_plant(paths["reports"], split_dir, "reports")
    size_files = _split_by_plant(paths["sizes"], split_dir, "sizes")
    assert len(report_files) > 1 and len(size_files) > 1, "нужно несколько заводов"

    def snapshot(source, sizes) -> dict:
        mapping = load_mapping()
        mapping.table_sizes.files = [str(p) for p in sizes]
        mapping.report_usage.files = []
        db = tmp_path_factory.mktemp("db") / "cmp.duckdb"
        build(source, db, mapping)
        con = duckdb.connect(str(db), read_only=True)
        one = lambda sql: con.execute(sql).fetchone()[0]  # noqa: E731
        out = {
            "size_rows": one("SELECT COUNT(*) FROM fact_table_size"),
            "sized_tables": one("SELECT COUNT(DISTINCT table_id) FROM fact_table_size"),
            "total_mb": round(one("SELECT COALESCE(SUM(total_mb), 0) FROM fact_table_size"), 1),
            "plants": one("SELECT COUNT(*) FROM (SELECT DISTINCT network, plant "
                          "FROM fact_table_size)"),
            "objects": one("SELECT COUNT(*) FROM dim_table"),
            "links": one("SELECT COUNT(*) FROM bridge_report_table"),
            "reports": one("SELECT COUNT(*) FROM dim_report WHERE plant IS NOT NULL"),
        }
        con.close()
        return out

    # Эталон — те же данные одним файлом. Строки без завода в нарезку не
    # попадают, поэтому отчёты сравниваем только те, у которых завод указан.
    assert snapshot(report_files, size_files) == snapshot(paths["reports"],
                                                          [paths["sizes"]])


def test_several_size_files_without_plant_are_reported(paths, tmp_path_factory):
    """Несколько файлов размеров без колонки завода — молча неверные цифры.

    Строки всех заводов лягут в «(не указан)» и сложатся в один размер, при
    этом никакой ошибки не возникнет. Единственная защита — предупреждение.
    """
    import pandas as pd

    split_dir = tmp_path_factory.mktemp("noplant")
    report_files = _split_by_plant(paths["reports"], split_dir, "reports")
    stripped = []
    for path in _split_by_plant(paths["sizes"], split_dir, "sizes"):
        target = split_dir / f"np_{path.name}"
        pd.read_excel(path, dtype=str).drop(columns=["ТС", "Завод"]).to_excel(
            target, index=False)
        stripped.append(target)

    mapping = load_mapping()
    mapping.table_sizes.files = [str(p) for p in stripped]
    mapping.report_usage.files = []
    db = tmp_path_factory.mktemp("db") / "noplant.duckdb"
    stats = build(report_files, db, mapping)

    assert stats.size_files_without_plant, "молчать об этом нельзя"

    # И предупреждение не напрасное: заводы действительно схлопнулись в один.
    con = duckdb.connect(str(db), read_only=True)
    plants = con.execute(
        "SELECT COUNT(*) FROM (SELECT DISTINCT network, plant FROM fact_table_size)"
    ).fetchone()[0]
    con.close()
    assert plants == 1


def test_one_size_file_without_plant_is_not_flagged(paths, tmp_path_factory):
    """Один файл без завода — обычный случай, тревожить не о чем."""
    import pandas as pd

    tmp = tmp_path_factory.mktemp("single")
    target = tmp / "sizes.xlsx"
    pd.read_excel(paths["sizes"], dtype=str).drop(columns=["ТС", "Завод"]).to_excel(
        target, index=False)

    mapping = load_mapping()
    mapping.table_sizes.files = [str(target)]
    mapping.report_usage.files = []
    stats = build(paths["reports"], tmp / "one.duckdb", mapping)
    assert not stats.size_files_without_plant


def test_section_config_accepts_one_file_or_many():
    """`file:` и `files:` в конфиге равноправны — старые конфиги работают."""
    single = SectionConfig.from_dict({"file": "a.xlsx"})
    many = SectionConfig.from_dict({"files": ["a.xlsx", "b.xlsx"]})
    assert single.files == ["a.xlsx"]
    assert single.file == "a.xlsx"
    assert many.files == ["a.xlsx", "b.xlsx"]
    assert many.file == "a.xlsx", "«первый файл» нужен коду, которому хватает одного"

    # Присваивание .file остаётся рабочим — им пользуются тесты и diagnose.
    single.file = "c.xlsx"
    assert single.files == ["c.xlsx"]
    single.file = None
    assert single.files == []


def test_schema_names_have_one_case(full_db):
    """Одна схема — одна запись, а не «dbo» и «DBO» по отдельности.

    Из отчётов схема приходит как «dbo», из выгрузки сегментов как «DBO».
    Без приведения к одному регистру сводки по схемам двоились: половина
    объёма уходила в схему-двойник.
    """
    total, lowered = full_db.execute(
        "SELECT COUNT(DISTINCT schema_name), COUNT(DISTINCT LOWER(schema_name)) "
        "FROM dim_table"
    ).fetchone()
    assert total == lowered


def test_schema_recovered_when_name_is_unique_in_sizes_file(full_db):
    """Схема без указания в отчёте восстанавливается по файлу размеров.

    Демо-данные ссылаются на «STANDALONE_1» без схемы; в файле размеров это
    имя существует только в схеме dbo. Должна получиться ровно одна запись
    dim_table с восстановленной схемой — а не дубль «(unknown).standalone_1»
    рядом с «dbo.standalone_1» из файла размеров.
    """
    rows = full_db.execute(
        "SELECT schema_name, schema_source, is_parsed_ok FROM dim_table "
        "WHERE table_name ILIKE 'standalone_1'"
    ).fetchall()
    assert len(rows) == 1, "восстановленная и файловая запись должны быть одной строкой"
    schema_name, schema_source, is_parsed_ok = rows[0]
    assert schema_name == "dbo"
    assert schema_source == "файл размеров"
    assert is_parsed_ok is True


def test_schema_not_guessed_when_name_is_ambiguous(full_db):
    """Несколько схем с одним именем — восстанавливать нельзя, это догадка.

    «Orders» в демо-данных существует сразу в нескольких схемах, поэтому
    ссылка без схемы должна остаться «(unknown)», а не взять первую попавшуюся.
    """
    row = full_db.execute(
        "SELECT schema_name, schema_source, is_parsed_ok FROM dim_table "
        "WHERE table_name ILIKE 'orders' AND schema_name = '(unknown)'"
    ).fetchone()
    assert row is not None, "в демо-данных есть ссылка на «Orders» без схемы"
    _, schema_source, is_parsed_ok = row
    assert schema_source == "не определена"
    assert is_parsed_ok is False


def test_is_parsed_ok_matches_schema_source(full_db):
    """is_parsed_ok — производная от schema_source, они не должны разойтись."""
    mismatched = full_db.execute(
        "SELECT COUNT(*) FROM dim_table "
        "WHERE is_parsed_ok <> (schema_source <> 'не определена')"
    ).fetchone()[0]
    assert mismatched == 0


def test_missing_segment_type_column_is_reported(paths, tmp_path_factory):
    """Без колонки типа индексы не отфильтровать — загрузка обязана сказать это.

    Самая дорогая из тихих ошибок: индексные сегменты становятся отдельными
    «таблицами» и завышают и число таблиц, и суммарный объём. Заказчик видит
    расхождение с исходным файлом и не понимает, откуда оно.
    """
    import pandas as pd

    df = pd.read_excel(paths["sizes"], dtype=str)
    broken = tmp_path_factory.mktemp("sizes") / "no_type.xlsx"
    df.drop(columns=["SEGMENT_TYPE"]).to_excel(broken, index=False)

    mapping = load_mapping()
    mapping.table_sizes.file = str(broken)
    db = tmp_path_factory.mktemp("db") / "no_type.duckdb"
    stats = build(paths["reports"], db, mapping)

    assert stats.segment_type_column_missing, "молчать об этом нельзя"

    # И убеждаемся, что предупреждение не напрасное: индексы действительно
    # просочились в список таблиц.
    con = duckdb.connect(str(db), read_only=True)
    leaked = con.execute(
        "SELECT COUNT(*) FROM dim_table WHERE table_name ILIKE 'IDX!_%' ESCAPE '!'"
    ).fetchone()[0]
    con.close()
    assert leaked > 0


def test_segment_type_column_present_is_not_flagged(full_db, paths):
    """На нормальном файле ложной тревоги быть не должно."""
    mapping = load_mapping()
    mapping.table_sizes.file = str(paths["sizes"])
    import tempfile

    with tempfile.TemporaryDirectory() as tmp:
        stats = build(paths["reports"], Path(tmp) / "ok.duckdb", mapping)
    assert not stats.segment_type_column_missing
    assert stats.segments_without_type == 0


def test_usage_comes_from_main_file(full_db):
    """Частота и длительность приходят из основного файла, без отдельного."""
    filled = full_db.execute(
        "SELECT COUNT(*) FROM fact_report_usage WHERE exec_count IS NOT NULL"
    ).fetchone()[0]
    assert filled > 0


def test_duration_is_seconds_not_milliseconds(full_db):
    """Длительность хранится в секундах — как в исходном файле."""
    worst = full_db.execute(
        "SELECT MAX(avg_duration_sec) FROM fact_report_usage"
    ).fetchone()[0]
    assert worst is not None and worst < 100000, (
        "значения похожи на миллисекунды — проверьте перевод единиц"
    )


def test_view_reports_get_low_confidence(full_db):
    """Отчёт через view не может получить высокую уверенность вывода."""
    bad = full_db.execute(
        "SELECT COUNT(*) FROM v_decommission_candidates "
        "WHERE uses_view AND confidence = 'Высокая'"
    ).fetchone()[0]
    assert bad == 0


def test_report_no_is_loaded(full_db):
    missing = full_db.execute(
        "SELECT COUNT(*) FROM dim_report WHERE report_no IS NULL"
    ).fetchone()[0]
    assert missing == 0


def test_raw_data_is_git_ignored():
    """Данные заказчика не должны попадать в репозиторий ни при каком раскладе."""
    ignore = (ROOT / ".gitignore").read_text(encoding="utf-8").splitlines()
    assert "data/raw/*" in ignore
    assert "data/*.duckdb" in ignore
    assert "dist/" in ignore


# --- Документация ---------------------------------------------------------

# Русский файл — оригинал, английский лежит рядом с суффиксом .en.
BILINGUAL_DOCS = [
    (ROOT / "README.md", ROOT / "README.en.md"),
    (ROOT / "CHANGELOG.md", ROOT / "CHANGELOG.en.md"),
    (ROOT / "docs" / "RUNBOOK.md", ROOT / "docs" / "RUNBOOK.en.md"),
    (ROOT / "docs" / "ARCHITECTURE.md", ROOT / "docs" / "ARCHITECTURE.en.md"),
    (ROOT / "docs" / "TZ.md", ROOT / "docs" / "TZ.en.md"),
]


def _headings(path: Path) -> list[int]:
    """Уровни заголовков документа: «##» → 2. Текст не сравниваем — он разный."""
    levels = []
    in_code = False
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.startswith("```"):
            in_code = not in_code
        elif not in_code and line.startswith("#"):
            levels.append(len(line) - len(line.lstrip("#")))
    return levels


@pytest.mark.parametrize("ru,en", BILINGUAL_DOCS, ids=lambda p: p.name)
def test_every_document_has_a_translation(ru, en):
    assert ru.exists(), f"нет русского оригинала: {ru}"
    assert en.exists(), f"нет английского перевода: {en}"


@pytest.mark.parametrize("ru,en", BILINGUAL_DOCS, ids=lambda p: p.name)
def test_translations_keep_the_same_structure(ru, en):
    """Одинаковый скелет заголовков у обеих версий.

    Без этой проверки версии расходятся за пару коммитов: правку внесли в
    русскую, английская осталась прежней и тихо превратилась во враньё.
    Сравниваются только уровни заголовков — текст, естественно, разный.
    """
    assert _headings(ru) == _headings(en), (
        f"структура {ru.name} и {en.name} разошлась: "
        f"{len(_headings(ru))} заголовков против {len(_headings(en))}"
    )


@pytest.mark.parametrize("ru,en", BILINGUAL_DOCS, ids=lambda p: p.name)
def test_documents_link_to_each_other(ru, en):
    """Переключатель языка есть в обоих файлах, иначе перевод не найти."""
    assert en.name in ru.read_text(encoding="utf-8")
    assert ru.name in en.read_text(encoding="utf-8")


def test_version_is_the_same_everywhere():
    """`pyproject.toml` и `config.VERSION` обязаны совпадать.

    VERSION пишется в `etl_run.tool_version`, то есть по любой базе видно,
    каким кодом она собрана. Если версии разъедутся, эта запись начнёт врать —
    а именно по ней разбираются, почему у двух людей разные цифры.
    """
    import re

    pyproject = (ROOT / "pyproject.toml").read_text(encoding="utf-8")
    declared = re.search(r'^version\s*=\s*"([^"]+)"', pyproject, re.M)
    assert declared, "в pyproject.toml нет version"

    from reportsdb.config import VERSION

    assert declared.group(1) == VERSION, (
        f"pyproject.toml обещает {declared.group(1)}, "
        f"а config.VERSION равен {VERSION}"
    )


def test_changelog_mentions_the_current_version():
    """Версия выпущена — значит, о ней написано в истории версий.

    Обе языковые версии: иначе английский CHANGELOG отстаёт молча.
    """
    from reportsdb.config import VERSION

    for name in ("CHANGELOG.md", "CHANGELOG.en.md"):
        text = (ROOT / name).read_text(encoding="utf-8")
        assert f"[{VERSION}]" in text, f"{name}: нет раздела про версию {VERSION}"


def test_decision_journal_is_kept_in_both_languages():
    """Журнал решений — часть документации, а не только русской версии."""
    count = lambda p: p.read_text(encoding="utf-8").count("\n| 2026-")  # noqa: E731
    ru = count(ROOT / "docs" / "TZ.md")
    en = count(ROOT / "docs" / "TZ.en.md")
    assert ru > 0 and ru == en, (
        f"записей в журнале: русских {ru}, английских {en} — "
        "новая запись добавляется в обе версии одним коммитом"
    )


# --- Целостность модели ---------------------------------------------------

def test_no_dangling_bridge_keys(full_db):
    dangling = full_db.execute(
        """
        SELECT COUNT(*) FROM bridge_report_table b
        LEFT JOIN dim_report r ON r.report_id = b.report_id
        LEFT JOIN dim_table  t ON t.table_id  = b.table_id
        WHERE r.report_id IS NULL OR t.table_id IS NULL
        """
    ).fetchone()[0]
    assert dangling == 0


def test_full_name_unique(full_db):
    dupes = full_db.execute(
        "SELECT COUNT(*) FROM (SELECT full_name FROM dim_table "
        "GROUP BY full_name HAVING COUNT(*) > 1)"
    ).fetchone()[0]
    assert dupes == 0


def test_schema_version_is_written(full_db):
    """Версия структуры пишется в журнал — по ней приложение узнаёт старую базу."""
    version = full_db.execute(
        "SELECT schema_version FROM etl_run ORDER BY run_id DESC LIMIT 1"
    ).fetchone()[0]
    assert version == SCHEMA_VERSION


def test_schema_version_matches_view_set():
    """Забыли поднять SCHEMA_VERSION после правки SQL — тест напомнит.

    Сверяем со слепком: любое изменение схемы или витрин обязано
    сопровождаться увеличением версии, иначе у пользователей останется
    несовместимая база и падение вместо понятного сообщения.
    """
    import hashlib

    digest = hashlib.sha256()
    for name in ("01_schema.sql", "02_views.sql"):
        digest.update((ROOT / "sql" / name).read_bytes())
    # Слепок SQL на момент SCHEMA_VERSION = 11.
    expected = "c5f87f"  # первые 6 знаков; обновлять вместе с версией
    actual = digest.hexdigest()[:6]
    assert actual == expected, (
        f"SQL изменился (слепок {actual}, ожидался {expected}). "
        f"Поднимите SCHEMA_VERSION в src/reportsdb/config.py и обновите слепок "
        f"в этом тесте."
    )


def test_report_overlap_stays_inside_one_plant(full_db):
    """Похожие отчёты ищутся только внутри одного завода.

    Один и тот же отчёт заведён на нескольких заводах — это нормальное
    устройство, а не дубль. Пары между заводами не только бесполезны, их ещё и
    столько, что за ними не видно настоящих кандидатов на объединение.
    """
    # Витрина несёт РЦ первого отчёта пары. Если второй на этом РЦ не
    # существует — значит пара межзаводская и просочилась мимо ограничения.
    cross = full_db.execute(
        """
        SELECT COUNT(*) FROM v_report_overlap o
        WHERE NOT EXISTS (
            SELECT 1 FROM dim_report r
            WHERE r.report_name = o.report_2
              AND COALESCE(r.network, '(не указана)') = o.network
              AND COALESCE(r.plant,   '(не указан)')  = o.plant
        )
        """
    ).fetchone()[0]
    assert cross == 0, "в витрину попала пара отчётов с разных заводов"


def test_export_overlap_query_stays_inside_one_plant(full_db):
    """То же ограничение в выгрузке для коллег — на живых строках.

    У витрины порог сходства 0.8, и на демо-данных она пуста, так что проверка
    выше срабатывает вхолостую. Запрос выгрузки идёт с порогом 0.3 и строки
    возвращает — на нём правило видно по-настоящему.
    """
    from reportsdb.export_html import QUERIES

    rows = full_db.execute(QUERIES["overlap"]).df()
    assert not rows.empty, "на демо-данных пары внутри завода обязаны находиться"

    plants = full_db.execute(
        "SELECT report_id, COALESCE(network, '') || '|' || COALESCE(plant, '') AS rc "
        "FROM dim_report"
    ).df().set_index("report_id")["rc"].to_dict()
    mixed = [(a, b) for a, b in zip(rows["id1"], rows["id2"])
             if plants[a] != plants[b]]
    assert not mixed, f"в выгрузку попали пары отчётов с разных заводов: {len(mixed)}"


# --- Уникальные отчёты завода ---------------------------------------------

def test_plant_twin_keeps_one_row_per_report(full_db):
    """Витрина двойников — строка на отчёт, а не на пару.

    Похожих отчётов у одного бывает несколько; если строки размножатся по их
    числу, суммы объёма и запусков на странице вырастут на пустом месте.
    """
    reports = full_db.execute("SELECT COUNT(*) FROM dim_report").fetchone()[0]
    rows = full_db.execute("SELECT COUNT(*) FROM v_report_plant_twin").fetchone()[0]
    assert rows == reports


def test_plant_twin_looks_only_at_other_plants_of_one_network(full_db):
    """Двойник ищется на ДРУГОМ заводе, но внутри своей ТС.

    Зеркальное правило к «Похожим отчётам»: там пара обязана быть внутри
    одного завода, здесь — обязана быть между разными заводами одной сети.
    Сети ведут хозяйство независимо, и совпадение отчёта между ними ничего не
    говорит о том, что завод делает сам.
    """
    bad = full_db.execute(
        """
        SELECT COUNT(*) FROM v_report_plant_twin v
        JOIN dim_report t ON t.report_id = v.best_twin_report_id
        JOIN dim_report r ON r.report_id = v.report_id
        WHERE COALESCE(t.network, '') <> COALESCE(r.network, '')
           OR COALESCE(t.plant, '')    = COALESCE(r.plant, '')
        """
    ).fetchone()[0]
    assert bad == 0, "двойник найден на своём заводе или в чужой ТС"


def _twin_db(tmp_path_factory):
    """Крошечная база с нарочно устроенными двойниками.

    На демо-данных наименования отчётов уникальны по построению, а сходство
    наборов таблиц не доходит до 0.3 — правила проверить нечем. Здесь каждый
    случай заведён вручную.
    """
    import pandas as pd

    rows = [
        # Тёзки в одной ТС: наборы таблиц разные, совпадает только имя —
        # и регистром с краевым пробелом, как в живой выгрузке.
        ("СЕТЬ-1", "Завод-А", "Отчёт-тёзка", "dbo.T1;dbo.T2;dbo.T3"),
        ("СЕТЬ-1", "Завод-Б", " отчёт-тёзка ", "dbo.T7;dbo.T8;dbo.T9"),
        # Разные имена, но набор таблиц совпадает полностью.
        ("СЕТЬ-1", "Завод-А", "Отчёт-копия-А", "dbo.T4;dbo.T5;dbo.T6"),
        ("СЕТЬ-1", "Завод-Б", "Отчёт-копия-Б", "dbo.T4;dbo.T5;dbo.T6"),
        # Своё и только своё.
        ("СЕТЬ-1", "Завод-А", "Отчёт-свой", "dbo.T10;dbo.T11"),
        # Одна общая таблица с «Отчёт-тёзка» завода А — это общий справочник,
        # а не тот же отчёт: порог «не меньше двух общих» обязан её отбросить.
        ("СЕТЬ-1", "Завод-Б", "Отчёт-со-справочником", "dbo.T1;dbo.T12"),
        # Полный двойник первого отчёта, но в ДРУГОЙ ТС — не в счёт.
        ("СЕТЬ-2", "Завод-В", "Отчёт-тёзка", "dbo.T1;dbo.T2;dbo.T3"),
    ]
    frame = pd.DataFrame(
        [
            {
                "№": i + 1,
                "ТС": network,
                "Завод": plant,
                "Каталог 1-го уровня": "Проверка",
                "Наименование отчета": name,
                "Таблицы источники данных": tables,
            }
            for i, (network, plant, name, tables) in enumerate(rows)
        ]
    )
    source = tmp_path_factory.mktemp("twin") / "twin_reports.xlsx"
    frame.to_excel(source, index=False)

    base = load_mapping()
    mapping = Mapping(
        reports=base.reports,
        table_sizes=SectionConfig(),
        report_usage=SectionConfig(),
    )
    db = tmp_path_factory.mktemp("twindb") / "twin.duckdb"
    build(source, db, mapping)
    return duckdb.connect(str(db), read_only=True)


@pytest.fixture(scope="module")
def twin_db(tmp_path_factory):
    con = _twin_db(tmp_path_factory)
    yield con
    con.close()


def test_plant_twin_catches_namesakes_across_plants(twin_db):
    """Тёзка на другом заводе своей ТС — отчёт не уникален.

    Сравнение по LOWER(TRIM(...)): регистр и краевые пробелы в выгрузке
    разъезжаются, а отчёт при этом один и тот же.
    """
    rows = twin_db.execute(
        "SELECT plant, name_twin_count, name_twin_plants FROM v_report_plant_twin "
        "WHERE LOWER(TRIM(report_name)) = 'отчёт-тёзка' AND network = 'СЕТЬ-1' "
        "ORDER BY plant"
    ).df()
    assert len(rows) == 2
    assert list(rows["name_twin_count"]) == [1, 1]
    assert rows.loc[0, "name_twin_plants"] == "Завод-Б"
    assert rows.loc[1, "name_twin_plants"] == "Завод-А"


def test_plant_twin_namesake_in_another_network_does_not_count(twin_db):
    """Тот же отчёт в другой ТС уникальности не отменяет.

    У отчёта из СЕТЬ-2 и имя, и набор таблиц совпадают с отчётом СЕТЬ-1 —
    и он всё равно обязан остаться уникальным для своего завода.
    """
    row = twin_db.execute(
        "SELECT name_twin_count, best_jaccard, plants_compared "
        "FROM v_report_plant_twin WHERE network = 'СЕТЬ-2'"
    ).fetchone()
    assert row[0] == 0, "тёзка найден в чужой ТС"
    assert row[1] is None, "двойник по таблицам найден в чужой ТС"
    assert row[2] == 0, "в СЕТЬ-2 один завод — сравнивать не с чем"


def test_plant_twin_finds_identical_table_sets(twin_db):
    """Полное совпадение набора таблиц — сходство 1.0 и двойник назван."""
    rows = twin_db.execute(
        "SELECT plant, best_jaccard, best_shared_tables, best_twin_report, "
        "best_twin_plant FROM v_report_plant_twin "
        "WHERE report_name LIKE 'Отчёт-копия%' ORDER BY plant"
    ).df()
    assert len(rows) == 2
    assert list(rows["best_jaccard"]) == [1.0, 1.0]
    assert list(rows["best_shared_tables"]) == [3, 3]
    assert list(rows["best_twin_plant"]) == ["Завод-Б", "Завод-А"]
    assert list(rows["best_twin_report"]) == ["Отчёт-копия-Б", "Отчёт-копия-А"]


def test_plant_twin_ignores_a_single_shared_table(twin_db):
    """Одна общая таблица — общий справочник, а не двойник.

    Без порога «не меньше двух общих» любой отчёт, читающий календарь, терял
    бы уникальность, и страница показывала бы нули там, где своего много.
    """
    row = twin_db.execute(
        "SELECT best_jaccard, name_twin_count, plants_compared "
        "FROM v_report_plant_twin WHERE report_name = 'Отчёт-со-справочником'"
    ).fetchone()
    assert row[0] is None, "пара на одной общей таблице просочилась в двойники"
    assert row[1] == 0
    assert row[2] == 1, "в СЕТЬ-1 есть второй завод — сравнивать было с чем"


def test_plant_twin_leaves_own_reports_unique(twin_db):
    """Отчёт без тёзки и без похожего набора таблиц остаётся уникальным."""
    row = twin_db.execute(
        "SELECT name_twin_count, best_jaccard, table_count "
        "FROM v_report_plant_twin WHERE report_name = 'Отчёт-свой'"
    ).fetchone()
    assert row == (0, None, 2)


def test_every_source_row_accounted_for(full_db):
    run = full_db.execute(
        "SELECT rows_read, rows_loaded, rows_rejected FROM etl_run"
    ).fetchone()
    assert run[0] == run[1] + run[2]


def test_exclusive_mb_never_exceeds_real_storage(full_db):
    """Ключевая защита от двойного учёта общих таблиц.

    Сравнение с полным объёмом всех заводов: отчёты живут на конкретных
    площадках, поэтому их сумма заведомо не больше.
    """
    total = full_db.execute("SELECT COALESCE(SUM(total_mb), 0) FROM fact_table_size").fetchone()[0]
    exclusive = full_db.execute(
        "SELECT COALESCE(SUM(exclusive_mb), 0) FROM v_report_footprint"
    ).fetchone()[0]
    assert exclusive <= total + 1e-6, (
        f"exclusive_mb={exclusive} превышает реальный объём {total} — "
        "значит, общие таблицы посчитаны дважды"
    )


def test_gross_mb_is_at_least_exclusive(full_db):
    bad = full_db.execute(
        "SELECT COUNT(*) FROM v_report_footprint WHERE gross_mb < exclusive_mb - 1e-6"
    ).fetchone()[0]
    assert bad == 0


def test_exclusive_tables_belong_to_exactly_one_report(full_db):
    """exclusive_mb должен строиться ровно на таблицах с report_count = 1."""
    mismatch = full_db.execute(
        """
        WITH per_report AS (
            SELECT b.report_id, COALESCE(SUM(s.total_mb), 0) AS mb
            FROM bridge_report_table b
            JOIN (SELECT table_id FROM bridge_report_table
                  GROUP BY table_id HAVING COUNT(DISTINCT report_id) = 1) e
              ON e.table_id = b.table_id
            -- Через резолвер: размер берётся для завода отчёта. Прямое
            -- соединение с fact_table_size размножило бы строки по заводам.
            LEFT JOIN v_report_table_size s
                   ON s.table_id = b.table_id AND s.report_id = b.report_id
            GROUP BY b.report_id
        )
        SELECT COUNT(*) FROM v_report_footprint f
        JOIN per_report p ON p.report_id = f.report_id
        WHERE ABS(p.mb - f.exclusive_mb) > 0.02
        """
    ).fetchone()[0]
    assert mismatch == 0


def test_size_coverage_within_bounds(full_db):
    bad = full_db.execute(
        "SELECT COUNT(*) FROM v_report_footprint "
        "WHERE size_coverage_pct IS NOT NULL "
        "AND (size_coverage_pct < 0 OR size_coverage_pct > 100)"
    ).fetchone()[0]
    assert bad == 0


def test_rejected_rows_are_logged(full_db):
    rejected = full_db.execute("SELECT rows_rejected FROM etl_run").fetchone()[0]
    logged = full_db.execute(
        "SELECT COUNT(*) FROM etl_reject WHERE reason = 'Пустое имя отчёта'"
    ).fetchone()[0]
    assert rejected == logged > 0, "строка с пустым именем должна попасть в etl_reject"


def test_report_without_sources_is_loaded(full_db):
    found = full_db.execute(
        "SELECT table_count FROM v_report_footprint WHERE report_name = 'Report Without Sources'"
    ).fetchone()
    assert found is not None and found[0] == 0


def test_every_loaded_field_is_visible_in_analytics(full_db):
    """Ни одно загруженное поле не должно остаться только в таблицах модели."""
    views = [
        row[0] for row in full_db.execute(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema = 'main' AND table_type = 'VIEW'"
        ).fetchall()
    ]
    exposed: set[str] = set()
    for view in views:
        exposed |= {row[0] for row in full_db.execute(f"DESCRIBE {view}").fetchall()}

    missing = [f for f in FIELDS_IN_ANALYTICS if f not in exposed]
    assert not missing, f"поля загружены, но не попали ни в одну витрину: {missing}"


def test_catalog_overview_keeps_all_three_levels(full_db):
    columns = {row[0] for row in full_db.execute("DESCRIBE v_catalog_overview").fetchall()}
    assert {"folder_l1", "folder_l2", "folder_l3"} <= columns
    deep = full_db.execute(
        "SELECT COUNT(*) FROM v_catalog_overview WHERE folder_l3 <> ''"
    ).fetchone()[0]
    assert deep > 0, "третий уровень каталога должен доходить до сводки"


def test_duration_bands_cover_all_reports_with_data(full_db):
    """У каждого отчёта с длительностью есть непустая категория."""
    bad = full_db.execute(
        "SELECT COUNT(*) FROM v_report_duration "
        "WHERE avg_duration_sec IS NOT NULL AND duration_band = 'Нет данных'"
    ).fetchone()[0]
    assert bad == 0


def test_total_duration_is_execs_times_average(full_db):
    mismatch = full_db.execute(
        """
        SELECT COUNT(*) FROM v_report_duration
        WHERE exec_count IS NOT NULL AND avg_duration_sec IS NOT NULL
          AND ABS(total_duration_sec - exec_count * avg_duration_sec) > 0.2
        """
    ).fetchone()[0]
    assert mismatch == 0


# --- Пять основных представлений в разрезе РЦ ----------------------------

def test_tables_catalog_is_exactly_the_size_file(full_db):
    """Таблица №1 — ровно строки файла размеров: ни потерь, ни добавок.

    Список таблиц БД берётся только из файла размеров. Объекты, упомянутые
    в отчётах, но отсутствующие в файле, в него не подмешиваются: мост
    «отчёт ↔ таблица» лишь ссылается на этот список.
    """
    sizes = full_db.execute("SELECT COUNT(*) FROM fact_table_size").fetchone()[0]
    in_catalog = full_db.execute("SELECT COUNT(*) FROM v_tables_catalog").fetchone()[0]
    assert in_catalog == sizes


def test_tables_catalog_adds_nothing_from_reports(full_db):
    """В каталоге нет ни одной таблицы, которой нет в файле размеров.

    В демо-данных отчёты заведомо ссылаются на объекты без замера (мусорные
    имена, view, временные) — они обязаны остаться за пределами №1.
    """
    dangling = full_db.execute(
        "SELECT COUNT(*) FROM dim_table t "
        "WHERE EXISTS (SELECT 1 FROM bridge_report_table b WHERE b.table_id = t.table_id) "
        "  AND NOT EXISTS (SELECT 1 FROM fact_table_size s WHERE s.table_id = t.table_id)"
    ).fetchone()[0]
    assert dangling > 0, "проверка бессмысленна, если таких ссылок нет вовсе"

    leaked = full_db.execute(
        "SELECT COUNT(*) FROM v_tables_catalog c "
        "WHERE NOT EXISTS (SELECT 1 FROM fact_table_size s JOIN dim_table t "
        "                  ON t.table_id = s.table_id WHERE t.full_name = c.full_name)"
    ).fetchone()[0]
    assert leaked == 0


def test_tables_catalog_is_independent_of_reports(full_db):
    """Таблица попадает в №1, даже если на неё не ссылается ни один отчёт."""
    orphans = full_db.execute(
        "SELECT COUNT(*) FROM v_tables_catalog WHERE report_count = 0"
    ).fetchone()[0]
    assert orphans > 0, "в демо-данных есть таблицы вне отчётов — они обязаны быть видны"


def test_tables_catalog_shows_nothing_twice(full_db):
    """Строка на пару «таблица + завод», без дублей."""
    dupes = full_db.execute(
        "SELECT COUNT(*) FROM (SELECT network, plant, full_name FROM v_tables_catalog "
        "GROUP BY 1, 2, 3 HAVING COUNT(*) > 1)"
    ).fetchone()[0]
    assert dupes == 0


def test_table_size_differs_per_plant(full_db):
    """Размер ведётся на завод: одна таблица на разных заводах весит своё."""
    varying = full_db.execute(
        "SELECT COUNT(*) FROM (SELECT table_id FROM fact_table_size "
        "GROUP BY table_id HAVING COUNT(DISTINCT plant) > 1 "
        "AND COUNT(DISTINCT total_mb) > 1)"
    ).fetchone()[0]
    assert varying > 0


def test_size_resolver_does_not_multiply_rows(full_db):
    """Резолвер обязан давать ровно одну строку размера на связь.

    Прямое соединение с fact_table_size по table_id размножило бы строки по
    числу заводов и завысило суммы в разы.
    """
    links = full_db.execute("SELECT COUNT(*) FROM bridge_report_table").fetchone()[0]
    resolved = full_db.execute("SELECT COUNT(*) FROM v_report_table_size").fetchone()[0]
    assert resolved == links


def test_size_resolver_picks_the_plant_of_the_report(full_db):
    """Отчёту достаётся размер его собственного завода, а не чужого."""
    wrong = full_db.execute(
        """
        SELECT COUNT(*) FROM v_report_table_size v
        JOIN dim_report r ON r.report_id = v.report_id
        JOIN fact_table_size s
          ON s.table_id = v.table_id
         AND s.network = COALESCE(r.network, '(не указана)')
         AND s.plant   = COALESCE(r.plant, '(не указан)')
        WHERE v.size_is_plant_specific AND v.total_mb IS DISTINCT FROM s.total_mb
        """
    ).fetchone()[0]
    assert wrong == 0


def test_rc_report_tables_contain_only_real_tables(full_db):
    """Таблица №2 — без view, mat.view, временных и процедур."""
    kinds = {
        row[0] for row in full_db.execute(
            "SELECT DISTINCT object_kind FROM v_rc_report_tables"
        ).fetchall()
    }
    assert kinds == {"TABLE"}, f"в таблицу №2 просочились: {kinds - {'TABLE'}}"


def test_rc_report_tables_is_one_to_many(full_db):
    """У отчёта несколько таблиц — связь не схлопнута в одну строку."""
    max_tables = full_db.execute(
        "SELECT MAX(n) FROM (SELECT COUNT(*) AS n FROM v_rc_report_tables "
        "GROUP BY network, plant, report_name)"
    ).fetchone()[0]
    assert max_tables > 1


def test_rc_routines_contain_only_routines(full_db):
    """Таблица №3 — только функции и процедуры."""
    kinds = {
        row[0] for row in full_db.execute(
            "SELECT DISTINCT t.object_kind FROM v_rc_report_routines v "
            "JOIN dim_table t ON t.full_name = v.routine_full_name"
        ).fetchall()
    }
    assert kinds == {"ROUTINE"}


def test_rc_usage_covers_every_report(full_db):
    """Таблица №4 — строка на каждый отчёт, даже без статистики."""
    reports = full_db.execute("SELECT COUNT(*) FROM dim_report").fetchone()[0]
    rows = full_db.execute("SELECT COUNT(*) FROM v_rc_report_usage").fetchone()[0]
    assert rows == reports


def test_rc_retention_bands_match_days(full_db):
    """Границы 30/45 расставлены ровно так, как просил заказчик."""
    bad = full_db.execute(
        """
        SELECT COUNT(*) FROM v_rc_report_retention
        WHERE (retention_days <= 30 AND retention_band <> 'До 30 дней')
           OR (retention_days > 30 AND retention_days <= 45
               AND retention_band <> 'От 31 до 45 дней')
           OR (retention_days > 45 AND retention_band <> 'Более 45 дней')
           OR (retention_days IS NULL AND retention_band <> 'Не задана')
        """
    ).fetchone()[0]
    assert bad == 0


def test_object_kinds_are_loaded_from_columns(full_db):
    """Типы объектов приходят из отдельных колонок, а не угадываются."""
    rows = dict(full_db.execute(
        "SELECT object_kind, COUNT(*) FROM dim_table GROUP BY 1"
    ).fetchall())
    for kind in ("TABLE", "VIEW", "MATERIALIZED VIEW", "TEMP", "ROUTINE"):
        assert rows.get(kind, 0) > 0, f"нет объектов типа {kind}"

    guessed = full_db.execute(
        "SELECT COUNT(*) FROM dim_table WHERE object_kind <> 'TABLE' "
        "AND kind_source <> 'колонка'"
    ).fetchone()[0]
    assert guessed == 0, "тип, отличный от таблицы, должен приходить из колонки"


def test_report_summary_totals_match_its_tables(full_db):
    """«Отчёт ссылается на такие-то таблицы, которые весят столько-то» —
    сумма в сводке обязана совпадать с суммой по её же таблицам."""
    mismatch = full_db.execute(
        """
        WITH per_report AS (
            SELECT b.report_id, ROUND(COALESCE(SUM(s.total_mb), 0), 2) AS mb,
                   COUNT(*) AS n
            FROM bridge_report_table b
            JOIN dim_table t ON t.table_id = b.table_id
            -- Только через резолвер: он даёт ровно одну строку размера на
            -- связь, для завода этого отчёта.
            LEFT JOIN v_report_table_size s
                   ON s.table_id = b.table_id AND s.report_id = b.report_id
            WHERE t.object_kind = 'TABLE'
            GROUP BY b.report_id
        )
        SELECT COUNT(*) FROM v_report_tables_summary v
        JOIN per_report p ON p.report_id = v.report_id
        WHERE ABS(p.mb - v.tables_total_mb) > 0.02 OR p.n <> v.table_count
        """
    ).fetchone()[0]
    assert mismatch == 0


def test_report_summary_lists_table_names(full_db):
    """В сводке перечислены сами таблицы, а не только их число."""
    row = full_db.execute(
        "SELECT table_count, table_names FROM v_report_tables_summary "
        "WHERE table_count > 1 LIMIT 1"
    ).fetchone()
    assert row and row[1] and row[1].count(",") == row[0] - 1


def test_report_summary_covers_every_report(full_db):
    reports = full_db.execute("SELECT COUNT(*) FROM dim_report").fetchone()[0]
    rows = full_db.execute("SELECT COUNT(*) FROM v_report_tables_summary").fetchone()[0]
    assert rows == reports


def test_only_the_view_prefix_mask_is_enabled():
    """Из масок имён включена ровно одна — префикс представления.

    Каждая маска — догадка о типе объекта, а неверно угаданный тип молча
    выкидывает таблицу из таблицы №2 и из расчёта объёма. Поэтому список
    закреплён тестом: новая маска добавляется осознанно, а не мимоходом.
    """
    assert load_mapping().reports.object_patterns == {"VIEW": ["V_*", "VW_*"]}


def test_view_prefix_recognised_but_never_overrides_real_tables(tmp_path):
    """Маска `v_` делает объект представлением — кроме тех, что есть в базе.

    Имя, начинающееся с «v_», почти всегда представление, и в справочнике его
    надо видеть отдельно от таблиц. Но если у объекта есть сегменты в файле
    размеров, значит он физически хранится, — это таблица, как бы она ни
    называлась, и маску к ней применять нельзя.
    """
    import pandas as pd

    reports = tmp_path / "reports.xlsx"
    pd.DataFrame([{
        "№": "1", "ТС": "СЕТЬ", "Завод": "З-1",
        "Каталог 1-го уровня": "Папка", "Наименование отчета": "Отчёт",
        # Все три объекта перечислены в колонке ТАБЛИЦ — то есть тип не задан
        # явно и решать его должна маска.
        "Таблицы источники данных": "dbo.orders; dbo.v_summary; dbo.v_stored",
    }]).to_excel(reports, index=False)

    sizes = tmp_path / "sizes.xlsx"
    pd.DataFrame([
        {"OWNER": "dbo", "SEGMENT_NAME": "orders", "SEGMENT_TYPE": "TABLE",
         "SIZE_MB": "10"},
        # Физически хранится, хотя и названо как представление.
        {"OWNER": "dbo", "SEGMENT_NAME": "v_stored", "SEGMENT_TYPE": "TABLE",
         "SIZE_MB": "20"},
    ]).to_excel(sizes, index=False)

    mapping = load_mapping()
    mapping.table_sizes.files = [str(sizes)]
    mapping.report_usage.files = []
    db = tmp_path / "kinds.duckdb"
    build(reports, db, mapping)

    con = duckdb.connect(str(db), read_only=True)
    kinds = dict(con.execute(
        "SELECT full_name, object_kind || '/' || kind_source FROM dim_table"
    ).fetchall())
    con.close()

    assert kinds["dbo.orders"].startswith("TABLE")
    assert kinds["dbo.v_summary"] == "VIEW/маска"
    assert kinds["dbo.v_stored"].startswith("TABLE"), (
        "объект с сегментами в файле размеров обязан остаться таблицей"
    )


# --- Поиск ----------------------------------------------------------------

@pytest.mark.parametrize(
    "text,expected",
    [
        ("dbo.orders", ["dbo.orders"]),
        ("dbo.orders, dbo.customers", ["dbo.orders", "dbo.customers"]),
        ("dbo.orders;dbo.customers", ["dbo.orders", "dbo.customers"]),
        ("dbo.orders\ndbo.customers", ["dbo.orders", "dbo.customers"]),
        ("  dbo.orders ,, ; dbo.customers  ", ["dbo.orders", "dbo.customers"]),
        # Пробел не разделитель: имя отчёта состоит из нескольких слов.
        ("Продажи за месяц", ["Продажи за месяц"]),
        ("", []),
        ("  , ; ", []),
    ],
)
def test_search_terms_splits_a_list_but_keeps_multiword_names(text, expected):
    from _shared import search_terms

    assert search_terms(text) == expected


def test_search_is_substring_not_regex():
    """Заказчик вставляет настоящие имена, а не регулярные выражения.

    `[dbo].[Orders]` как регулярка — это класс символов, он нашёл бы строки с
    буквами d, b, o; одинокая `*` уронила бы страницу «nothing to repeat».
    """
    import pandas as pd

    from _shared import search_terms

    assert search_terms("[dbo].[Orders]") == ["[dbo].[Orders]"]

    frame = pd.DataFrame({"full_name": ["[dbo].[Orders]", "fin.invoice"]})
    hit = frame["full_name"].str.contains("[dbo].[Orders]", case=False,
                                          na=False, regex=False)
    assert list(hit) == [True, False]


def test_exact_match_finds_only_the_listed_names():
    """Список готовых имён не должен приносить соседей по подстроке.

    Ровно та жалоба заказчика: он вставил список таблиц, а в выборку попали
    `dbo.catalog` и `wh1.errlog` — просто потому, что «log» есть и в них.
    """
    import pandas as pd

    names = pd.Series([
        "itm.log", "dbo.catalog", "wh1.errlog", "wh1.transmitlog",
        "sdd.trip_log", "fin.invoice",
    ])

    # Как ищет режим «часть имени»: соседи приезжают вместе с нужной строкой.
    loose = names[names.str.contains("itm.log", case=False, regex=False)]
    assert list(loose) == ["itm.log"]
    loose_log = names[names.str.contains("log", case=False, regex=False)]
    assert len(loose_log) == 5, "по части имени «log» находится половина базы"

    # Как ищет режим «точное совпадение».
    lowered = names.str.lower()
    assert list(names[lowered.eq("itm.log")]) == ["itm.log"]

    # Имя без схемы тоже совпадение: в выгрузке половина имён без неё.
    bare = lowered.str.rsplit(".", n=1).str[-1]
    assert list(names[lowered.eq("trip_log") | bare.eq("trip_log")]) == ["sdd.trip_log"]


# --- Выгрузка в Excel -----------------------------------------------------

def test_workbook_has_every_sheet_with_russian_headers():
    """Книгу открывают в Excel люди, а не программа.

    Значит: листов столько, сколько попросили; шапка по-русски, потому что
    `table_full_name` читателю ничего не говорит; шапка закреплена, иначе на
    длинном списке непонятно, где какая колонка.
    """
    from io import BytesIO

    import openpyxl
    import pandas as pd

    from _shared import build_workbook

    pairs = pd.DataFrame({
        "table_full_name": ["dbo.category", "dbo.category", "fin.invoice"],
        "report_name": ["Отчёт A", "Отчёт Б", "Отчёт A"],
        "total_mb": [10.0, 10.0, 5.0],
        # Технический ключ обязан выпасть: в Excel он только мешает.
        "table_id": [1, 1, 2],
    })
    by_report = pd.DataFrame({
        "report_name": ["Отчёт A", "Отчёт Б"],
        "table_count": [2, 1],
        "table_names": ["dbo.category; fin.invoice", "dbo.category"],
    })

    book = openpyxl.load_workbook(BytesIO(build_workbook({
        "Отчёты": by_report, "Таблицы и отчёты": pairs,
    })))
    assert book.sheetnames == ["Отчёты", "Таблицы и отчёты"]

    head = [c.value for c in book["Отчёты"][1]]
    assert head == ["Отчёт", "Таблиц", "Таблицы"]
    assert book["Отчёты"].max_row == 3, "две строки данных плюс шапка"
    assert book["Отчёты"].freeze_panes == "A2"

    pair_head = [c.value for c in book["Таблицы и отчёты"][1]]
    assert "table_id" not in pair_head, "суррогатный ключ в файл не идёт"
    assert pair_head == ["Таблица", "Отчёт", "Объём, МБ"]


def test_workbook_sheet_name_survives_excel_limits():
    """Excel не принимает имя листа длиннее 31 знака и символы : \\ / ? * [ ]."""
    from io import BytesIO

    import openpyxl
    import pandas as pd

    from _shared import build_workbook

    book = openpyxl.load_workbook(BytesIO(build_workbook({
        "Отчёты: таблицы/размеры [весь список за период]": pd.DataFrame({"a": [1]}),
    })))
    name = book.sheetnames[0]
    assert len(name) <= 31
    assert not set(name) & set(':\\/?*[]')


# --- Образ для передачи коллегам ------------------------------------------

def test_dockerfile_copies_everything_the_app_needs():
    """Всё, что приложение читает во время работы, обязано попасть в образ.

    Забытый каталог не роняет сборку — он тихо меняет поведение у коллеги.
    Так однажды потерялся `.streamlit/`: в контейнере пропадала тёмная тема,
    `backgroundColor` переставал совпадать с `SURFACE` (кайма вокруг сегментов
    диаграмм) и включалась обратно отправка статистики использования.
    """
    dockerfile = (ROOT / "docker" / "Dockerfile").read_text(encoding="utf-8")
    for folder in ("src/", "app/", "sql/", "config/", "data/", ".streamlit/"):
        assert f"COPY {folder}" in dockerfile, f"Dockerfile не копирует {folder}"


def test_dockerignore_keeps_the_data_folder():
    """`data/` намеренно попадает в образ — коллеге хватает одной команды.

    Если каталог однажды окажется в `.dockerignore`, образ соберётся молча, а
    у коллеги приложение встретит его пустой базой.
    """
    ignored = (ROOT / ".dockerignore").read_text(encoding="utf-8").split()
    assert "data" not in ignored and "data/" not in ignored


# --- Очистка базы ---------------------------------------------------------

def test_clear_leaves_an_empty_but_working_database(paths, tmp_path):
    """Очистка стирает данные, но оставляет базу, на которой работает аналитика.

    Пустая база, а не удалённый файл: иначе каждая страница встречала бы
    пользователя ошибкой «база не найдена» вместо честных нулей.
    """
    mapping = load_mapping()
    mapping.table_sizes.file = str(paths["sizes"])
    mapping.report_usage.file = str(paths["usage"])
    db = tmp_path / "reports.duckdb"
    build(paths["reports"], db, mapping)

    con = duckdb.connect(str(db), read_only=True)
    assert con.execute("SELECT COUNT(*) FROM dim_report").fetchone()[0] > 0
    con.close()

    result = clear(db)
    assert db.exists(), "файл базы должен остаться — пустым, но рабочим"
    assert result.freed_bytes > 0

    con = duckdb.connect(str(db), read_only=True)
    for table in ("dim_report", "dim_table", "bridge_report_table",
                  "fact_table_size", "fact_report_usage", "etl_reject"):
        assert con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0] == 0, table

    # Витрины обязаны работать на пустой базе — иначе очистка ломает приложение.
    for view in VIEWS:
        con.execute(f"SELECT * FROM {view}").fetchall()

    # По журналу видно, что базу очистили, а не что её никогда не собирали.
    run = con.execute(
        "SELECT source_file, rows_read, schema_version FROM etl_run"
    ).fetchone()
    con.close()
    assert run[1] == 0
    assert run[2] == SCHEMA_VERSION, "версия структуры должна остаться текущей"


def test_clear_removes_the_backup_unless_asked_to_keep_it(paths, tmp_path):
    """По умолчанию `.bak` удаляется вместе с базой.

    Очистку просят, когда рабочих данных на машине быть не должно. Копия
    рядом с пустой базой сохранила бы ровно то, что просили стереть.
    """
    mapping = load_mapping()
    mapping.table_sizes.file = str(paths["sizes"])
    mapping.report_usage.file = str(paths["usage"])

    for keep, must_exist in ((False, False), (True, True)):
        db = tmp_path / f"keep_{keep}.duckdb"
        build(paths["reports"], db, mapping)
        build(paths["reports"], db, mapping)  # вторая сборка создаёт .bak
        backup = db.with_suffix(db.suffix + ".bak")
        assert backup.exists(), "вторая сборка обязана оставить резервную копию"

        result = clear(db, keep_backup=keep)
        assert backup.exists() is must_exist
        assert result.backup_removed is not keep


def test_clear_does_not_touch_source_files(paths, tmp_path):
    """Исходные файлы пользователя — не содержимое базы, их очистка не трогает."""
    mapping = load_mapping()
    mapping.table_sizes.file = str(paths["sizes"])
    db = tmp_path / "reports.duckdb"
    build(paths["reports"], db, mapping)

    clear(db)
    assert paths["reports"].exists()
    assert paths["sizes"].exists()


# --- Витрины на пустых фактах --------------------------------------------

@pytest.mark.parametrize("view", VIEWS)
def test_views_work_without_facts(bare_db, view):
    """Главное требование ТЗ: аналитика не падает до появления размеров и статистики."""
    bare_db.execute(f"SELECT * FROM {view}").fetchall()


def test_footprint_is_zero_without_sizes(bare_db):
    total = bare_db.execute("SELECT SUM(exclusive_mb) FROM v_report_footprint").fetchone()[0]
    assert total == 0


def test_footprint_survives_without_size_file(bare_db):
    """Без файла размеров объёмы нулевые, но витрины считаются."""
    rows = bare_db.execute(
        "SELECT COUNT(*), COUNT(*) FILTER (WHERE exclusive_mb = 0) FROM v_report_footprint"
    ).fetchone()
    assert rows[0] == rows[1] > 0


def test_quadrants_use_usage_from_main_file(bare_db):
    """Статистика лежит в основном файле, поэтому квадранты считаются и без
    отдельного файла статистики."""
    quadrants = {
        row[0] for row in bare_db.execute(
            "SELECT DISTINCT quadrant FROM v_report_cost_value"
        ).fetchall()
    }
    assert quadrants - {"Нет данных об использовании"}


# --- Экспорт HTML ---------------------------------------------------------

def test_html_export_is_self_contained(full_db, tmp_path, paths):
    import re

    from reportsdb.export_html import export

    mapping = load_mapping()
    mapping.table_sizes.file = str(paths["sizes"])
    mapping.report_usage.file = str(paths["usage"])
    db = tmp_path / "export.duckdb"
    build(paths["reports"], db, mapping)

    out = export(db, tmp_path / "out.html")
    html = out.read_text(encoding="utf-8")

    assert "/*__DATA__*/null" not in html, "данные не подставились в шаблон"
    # Организационный разрез и время должны доехать до автономного файла.
    for key in ("network", "plant", "uses_view", "durations", "duration_band"):
        assert key in html, f"в HTML не попало поле {key}"
    external = re.findall(r"""(?:src|href)\s*=\s*["']https?://""", html)
    assert not external, f"HTML тянет внешние ресурсы: {external}"
    assert "Отчётность SSRS" in html
